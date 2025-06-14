from datetime import datetime
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import NamedStyle, Font, Alignment
from openpyxl.utils import get_column_letter

def export_logs_to_excel(logs):
    output = BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = "Time Logs"

    headers = [
        ("Staff", 20),
        ("Barcode", 15),
        ("Item Name", 30),
        ("Quantity", 10),
        ("Type", 10),
        ("Timestamp", 20)
    ]

    # Write headers with styling
    header_font = Font(bold=True)
    for col_idx, (header, width) in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Create a style that forces dd/mm/yyyy format
    if "date_style" not in wb.named_styles:
        date_style = NamedStyle(name="date_style", number_format="DD/MM/YYYY HH:MM:SS")
        wb.add_named_style(date_style)

    for log in logs:
        qty = log.get("qty_change", 0)
        log_type = "Masuk" if qty > 0 else "Keluar"
        raw_timestamp = log.get("timestamp", "")

        # Parse timestamp and format as dd/mm/yyyy string
        try:
            dt = datetime.strptime(raw_timestamp, "%Y-%m-%d %H:%M:%S")
            formatted_timestamp = dt.strftime("%d/%m/%Y %H:%M:%S")
            is_valid_date = True
        except (ValueError, TypeError):
            formatted_timestamp = raw_timestamp  # fallback to original if parsing fails
            is_valid_date = False

        row = [
            log.get("username", ""),
            log.get("barcode", ""),
            log.get("name", ""),
            abs(qty),
            log_type,
            formatted_timestamp
        ]
        ws.append(row)

        # Format the timestamp cell
        timestamp_cell = ws.cell(row=ws.max_row, column=6)
        if is_valid_date:
            timestamp_cell.style = "date_style"
        else:
            timestamp_cell.number_format = '@'  # Text format for invalid dates

    # Add filters and freeze header
    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"

    wb.save(output)
    output.seek(0)
    return output